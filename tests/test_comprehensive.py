"""
Comprehensive cross-validation: openpyxl, xlsxwriter, xlspy, Excel COM.
All temp files created dynamically, cleaned up after tests.
Tests performed on Windows 11.
"""
import os, sys, tempfile, datetime, shutil
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../src'))
from xlspy import XlsbWriter, XlsxWriter, ExcelReader
from conftest import PROJECT_TMP

# ── Diverse test data ───────────────────────────────────────────────────
HEADER = ["strings", "ints", "floats", "bools", "dates", "datetimes", "mixed", "empty"]

ROWS = [
    ["Alice",   42,               3.14159,  True,  datetime.date(2024,1,15),  datetime.datetime(2024,1,15,10,30,0), "hello", None],
    ["Bob",     0,               -0.001,   False, datetime.date(2023,6,1),   datetime.datetime(2023,6,1,0,0,0),    123,     None],
    ["Charlie", -1,               1e10,     True,  datetime.date(2020,12,31), datetime.datetime(2020,12,31,23,59,59), 3.14, None],
    [" ",       2147483647,       0.0000001,False, datetime.date(1900,1,1),  datetime.datetime(1900,1,1,0,0,0),    True,    None],
    ["Zazolc",  -2147483648,      1.23456789e-8, True,  datetime.date(1999,9,9), datetime.datetime(1999,9,9,9,9,9),None,    None],
]

ALL_DATA = [HEADER] + ROWS

# ── Fixture: temp dir inside project root (visible, then cleaned) ──────
@pytest.fixture(scope="module")
def tmp_dir():
    os.makedirs(PROJECT_TMP, exist_ok=True)
    d = tempfile.mkdtemp(prefix="pyxlsb_test_", dir=PROJECT_TMP)
    yield d
    shutil.rmtree(d, ignore_errors=True)


# ── Helpers ────────────────────────────────────────────────────────────
def _write_openpyxl(tmp, data):
    import openpyxl
    from openpyxl.styles import Font
    path = os.path.join(tmp, "from_openpyxl.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    for r, row in enumerate(data, 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if r == 1: cell.font = Font(bold=True)
    wb.save(path); wb.close()
    return path

def _write_xlsxwriter(tmp, data):
    import xlsxwriter as xw
    path = os.path.join(tmp, "from_xlsxwriter.xlsx")
    wb = xw.Workbook(path)
    ws = wb.add_worksheet()
    bold = wb.add_format({'bold': True})
    date_fmt = wb.add_format({'num_format': 'yyyy-mm-dd'})
    dt_fmt  = wb.add_format({'num_format': 'yyyy-mm-dd hh:mm:ss'})
    for c, h in enumerate(data[0]):
        ws.write(0, c, h, bold)
    for r, row in enumerate(data[1:], 1):
        for c, val in enumerate(row):
            if isinstance(val, datetime.datetime):
                ws.write_datetime(r, c, val, dt_fmt)
            elif isinstance(val, datetime.date):
                ws.write_datetime(r, c, val, date_fmt)
            elif val is None:
                ws.write_blank(r, c, None)
            else:
                ws.write(r, c, val)
    wb.close()
    return path

def _write_our_xlsx(tmp, data):
    path = os.path.join(tmp, "from_pyxlsb.xlsx")
    with XlsxWriter(path) as w:
        w.add_sheet("Sheet1"); w.write_sheet(data)
    return path

def _write_our_xlsb(tmp, data):
    path = os.path.join(tmp, "from_pyxlsb.xlsb")
    with XlsbWriter(path) as w:
        w.add_sheet("Sheet1"); w.write_sheet(data)
    return path

def _read_all(path):
    r = ExcelReader(); r.open(path)
    rows = list(r.get_rows(r.get_sheet_names()[0]))
    r.close()
    return rows

def _assert_rows_match(got, expected, label):
    assert len(got) == len(expected), f"{label}: row cnt {len(got)} vs {len(expected)}"
    for ri, (g, e) in enumerate(zip(got, expected)):
        assert len(g) == len(e), f"{label} row {ri}: col cnt {len(g)} vs {len(e)}"
        for ci, (gv, ev) in enumerate(zip(g, e)):
            if ev is None:
                assert gv is None or gv == '', f"{label} [{ri},{ci}]: expected None/empty, got {gv!r}"
            elif isinstance(ev, bool):
                assert gv == ev, f"{label} [{ri},{ci}]: bool {ev} vs {gv!r}"
            elif isinstance(ev, int):
                if isinstance(gv, float):
                    assert abs(gv - ev) < 1e-9, f"{label} [{ri},{ci}]: int {ev} vs {gv!r}"
                else:
                    assert gv == ev, f"{label} [{ri},{ci}]: int {ev} vs {gv!r}"
            elif isinstance(ev, float):
                assert isinstance(gv, (int, float)), f"{label} [{ri},{ci}]: float {ev} vs {type(gv).__name__} {gv!r}"
                assert abs(gv - ev) / max(1, abs(ev)) < 1e-5, f"{label} [{ri},{ci}]: float {ev} vs {gv!r}"
            elif isinstance(ev, (datetime.datetime, datetime.date)):
                assert isinstance(gv, (datetime.datetime, datetime.date)), f"{label} [{ri},{ci}]: date {ev} vs {type(gv).__name__} {gv!r}"
            else:
                assert gv == ev, f"{label} [{ri},{ci}]: {ev!r} vs {gv!r}"


# ─── TESTS ─────────────────────────────────────────────────────────────

class TestCrossValidation:

    # ─── Cross-format reads ──────────────────────────────────────────
    def test_read_openpyxl(self, tmp_dir):
        _assert_rows_match(_read_all(_write_openpyxl(tmp_dir, ALL_DATA)), ALL_DATA, "openpyxl->pyxlsb")

    def test_read_xlsxwriter(self, tmp_dir):
        _assert_rows_match(_read_all(_write_xlsxwriter(tmp_dir, ALL_DATA)), ALL_DATA, "xlsxwriter->pyxlsb")

    def test_read_our_xlsx(self, tmp_dir):
        _assert_rows_match(_read_all(_write_our_xlsx(tmp_dir, ALL_DATA)), ALL_DATA, "our XLSX->pyxlsb")

    def test_read_our_xlsb(self, tmp_dir):
        _assert_rows_match(_read_all(_write_our_xlsb(tmp_dir, ALL_DATA)), ALL_DATA, "our XLSB->pyxlsb")

    # ─── Other libs read our files ───────────────────────────────────
    def test_openpyxl_reads_our_xlsx(self, tmp_dir):
        import openpyxl
        path = _write_our_xlsx(tmp_dir, ALL_DATA)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        rows = [list(row) for row in wb.active.iter_rows(values_only=True)]
        wb.close()
        # openpyxl may return None for empty strings; we accept both
        for ri, (g, e) in enumerate(zip(rows, ALL_DATA)):
            for ci, (gv, ev) in enumerate(zip(g, e)):
                if ev is None:
                    assert gv is None or gv == '', f"openpyxl [{ri},{ci}]: {gv!r}"
                elif isinstance(ev, (datetime.date,)):
                    assert isinstance(gv, (datetime.datetime, datetime.date)), f"openpyxl [{ri},{ci}]: {gv!r}"
                elif isinstance(ev, bool):
                    assert gv == ev, f"openpyxl [{ri},{ci}]: {gv!r}"

    def test_our_xlsb_is_valid_zip(self, tmp_dir):
        import zipfile
        with zipfile.ZipFile(_write_our_xlsb(tmp_dir, ALL_DATA)) as z:
            assert len(z.namelist()) > 5, "Too few entries in XLSB"

    # ─── Roundtrip ───────────────────────────────────────────────────
    def test_roundtrip_xlsb_to_xlsx(self, tmp_dir):
        path = os.path.join(tmp_dir, "rt.xlsx")
        with XlsxWriter(path) as w:
            w.add_sheet("S"); w.write_sheet(_read_all(_write_our_xlsb(tmp_dir, ALL_DATA)))
        assert _read_all(path) == _read_all(_write_our_xlsb(tmp_dir, ALL_DATA))

    def test_roundtrip_xlsx_to_xlsb(self, tmp_dir):
        path = os.path.join(tmp_dir, "rt.xlsb")
        with XlsbWriter(path) as w:
            w.add_sheet("S"); w.write_sheet(_read_all(_write_our_xlsx(tmp_dir, ALL_DATA)))
        assert _read_all(path) == _read_all(_write_our_xlsx(tmp_dir, ALL_DATA))

    # ─── Shared string uniqueness ────────────────────────────────────
    def test_shared_string_uniqueness(self, tmp_dir):
        import string as s
        chars = s.ascii_letters + s.digits
        n = 1000
        data = []
        for i in range(1, n + 1):
            x = i; buf = []
            while x > 0:
                buf.append(chars[x % len(chars)])
                x //= len(chars)
            sval = ''.join(reversed(buf)).rjust(24, chars[0])[:24]
            data.append([sval])
        path = os.path.join(tmp_dir, "ss.xlsb")
        with XlsbWriter(path) as w:
            w.add_sheet("S"); w.write_sheet(data)
        seen = set()
        for r in _read_all(path):
            seen.add(r[0])
        assert len(seen) == n, f"{n - len(seen)} duplicate shared strings"

    # ─── Edge cases ──────────────────────────────────────────────────
    def test_basic_edge_cases(self, tmp_dir):
        cases = [
            [["A"]],
            [[1], [2], [3]],
            [[-2147483648, 2147483647, 0]],
            [[True], [False]],
        ]
        for idx, data in enumerate(cases):
            path = os.path.join(tmp_dir, f"e{idx}.xlsb")
            with XlsbWriter(path) as w:
                w.add_sheet("S"); w.write_sheet(data)
            rows = _read_all(path)
            assert rows == data, f"Edge case {idx}: {rows} != {data}"

    def test_none_row(self, tmp_dir):
        """All-None row only in our own roundtrip (xlsxwriter/Excel skip such rows)."""
        data = [["A", "B"], [None, None], ["C", "D"]]
        path = os.path.join(tmp_dir, "none_row.xlsb")
        with XlsbWriter(path) as w:
            w.add_sheet("S"); w.write_sheet(data)
        rows = _read_all(path)
        assert rows == data, f"None row: {rows} != {data}"

    # ─── Excel COM (conditional on win32com) ─────────────────────────
    def test_excel_opens_our_xlsb(self, tmp_dir):
        pytest.importorskip("win32com.client")
        import win32com.client
        path = _write_our_xlsb(tmp_dir, ALL_DATA)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False; excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Open(os.path.abspath(path))
            ws = wb.ActiveSheet
            assert ws.UsedRange.Rows.Count == len(ALL_DATA)
            assert ws.UsedRange.Columns.Count == len(HEADER)
            for ci, h in enumerate(HEADER, 1):
                assert ws.Cells(1, ci).Value == h, f"header col {ci}: {ws.Cells(1, ci).Value!r}"
            assert ws.Cells(2, 1).Value == "Alice"
            wb.Close(False)
        finally:
            excel.Quit()

    def test_excel_opens_our_xlsx(self, tmp_dir):
        pytest.importorskip("win32com.client")
        import win32com.client
        path = _write_our_xlsx(tmp_dir, ALL_DATA)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False; excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Open(os.path.abspath(path))
            ws = wb.ActiveSheet
            assert ws.Cells(1, 1).Value == "strings"
            assert ws.Cells(2, 1).Value == "Alice"
            wb.Close(False)
        finally:
            excel.Quit()

    def test_excel_write_and_readback(self, tmp_dir):
        pytest.importorskip("win32com.client")
        import win32com.client
        data = [["Tool", "Value"], ["xlspy", 123], ["Excel", 456.789]]
        path = os.path.join(tmp_dir, "ew.xlsb")
        with XlsbWriter(path) as w:
            w.add_sheet("T"); w.write_sheet(data)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False; excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Open(os.path.abspath(path))
            ws = wb.ActiveSheet
            assert ws.Cells(1, 1).Value == "Tool"
            assert ws.Cells(2, 2).Value == 123
            assert abs(ws.Cells(3, 2).Value - 456.789) < 0.001
            wb.Close(False)
        finally:
            excel.Quit()
